"""
Report generation module for SpecGuard.

Generates audit-ready output artifacts:
- Traceability matrix (CSV, optionally XLSX)
- Compliance report (Markdown, optionally PDF)
- JSON exports for programmatic access
"""

import csv
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from .models import (
    AnalysisResult,
    AnalysisSummary,
    ComplianceResult,
    GeneratedArtifacts,
    Requirement,
    RiskLevel,
    Verdict,
)

logger = logging.getLogger(__name__)


def generate_artifacts(
    requirements: list[Requirement],
    results: list[ComplianceResult],
    spec_filename: str,
    design_filename: str,
    output_dir: Path,
    analysis_id: str
) -> GeneratedArtifacts:
    """
    Generate all output artifacts for an analysis.

    Args:
        requirements: Extracted requirements
        results: Compliance check results
        spec_filename: Name of specification file
        design_filename: Name of design file
        output_dir: Directory to write outputs
        analysis_id: Unique identifier for this analysis

    Returns:
        GeneratedArtifacts with paths to all generated files
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    # Build requirement lookup
    req_lookup = {req.id: req for req in requirements}

    # Generate each artifact
    csv_path = generate_traceability_matrix(
        requirements, results, output_dir, analysis_id
    )

    md_path = generate_compliance_report(
        requirements, results, spec_filename, design_filename, output_dir, analysis_id
    )

    req_json_path = generate_requirements_json(
        requirements, output_dir, analysis_id
    )

    full_json_path = generate_full_results_json(
        requirements, results, spec_filename, design_filename, output_dir, analysis_id
    )

    return GeneratedArtifacts(
        traceability_matrix_csv=str(csv_path),
        compliance_report_md=str(md_path),
        requirements_json=str(req_json_path),
        full_results_json=str(full_json_path)
    )


def generate_traceability_matrix(
    requirements: list[Requirement],
    results: list[ComplianceResult],
    output_dir: Path,
    analysis_id: str
) -> Path:
    """
    Generate traceability matrix CSV.

    Columns: requirement_id, clause, requirement, verdict, evidence_ref, gap_summary, risk
    """
    result_lookup = {r.requirement_id: r for r in results}
    output_path = output_dir / f"traceability_matrix_{analysis_id}.csv"

    logger.info(f"Generating traceability matrix: {output_path}")

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        # Header
        writer.writerow([
            "requirement_id",
            "clause",
            "requirement",
            "type",
            "priority",
            "verdict",
            "evidence_ref",
            "gap_summary",
            "risk",
            "risk_reason"
        ])

        for req in requirements:
            result = result_lookup.get(req.id)

            if result:
                # Format evidence references
                evidence_refs = []
                for ev in result.evidence[:3]:  # Limit to top 3
                    loc = str(ev.source_location)
                    evidence_refs.append(loc)
                evidence_ref = "; ".join(evidence_refs) if evidence_refs else "No evidence"

                # Format gaps
                gap_summary = "; ".join(
                    gap.description for gap in result.gaps
                ) if result.gaps else ""

                verdict = result.verdict.value
                risk = result.risk.value
                risk_reason = result.risk_reason
            else:
                evidence_ref = "Not evaluated"
                gap_summary = ""
                verdict = "Not evaluated"
                risk = ""
                risk_reason = ""

            writer.writerow([
                req.id,
                req.clause or "",
                req.description[:200],  # Truncate long descriptions
                req.type.value,
                req.priority.value,
                verdict,
                evidence_ref,
                gap_summary,
                risk,
                risk_reason
            ])

    logger.info(f"Traceability matrix generated: {len(requirements)} requirements")
    return output_path


def generate_compliance_report(
    requirements: list[Requirement],
    results: list[ComplianceResult],
    spec_filename: str,
    design_filename: str,
    output_dir: Path,
    analysis_id: str
) -> Path:
    """
    Generate compliance report in Markdown format.

    Includes executive summary, detailed findings, and recommendations.
    """
    result_lookup = {r.requirement_id: r for r in results}
    output_path = output_dir / f"compliance_report_{analysis_id}.md"

    logger.info(f"Generating compliance report: {output_path}")

    # Calculate statistics
    total = len(requirements)
    compliant = sum(1 for r in results if r.verdict == Verdict.COMPLIANT)
    partial = sum(1 for r in results if r.verdict == Verdict.PARTIAL)
    missing = sum(1 for r in results if r.verdict == Verdict.MISSING)
    unknown = sum(1 for r in results if r.verdict == Verdict.UNKNOWN)

    high_risk = sum(1 for r in results if r.risk == RiskLevel.HIGH)
    medium_risk = sum(1 for r in results if r.risk == RiskLevel.MEDIUM)
    low_risk = sum(1 for r in results if r.risk == RiskLevel.LOW)

    compliance_rate = (compliant / total * 100) if total > 0 else 0

    with open(output_path, "w", encoding="utf-8") as f:
        # Header
        f.write(f"# Compliance Report\n\n")
        f.write(f"**Analysis ID:** {analysis_id}\n\n")
        f.write(f"**Generated:** {datetime.utcnow().isoformat()}Z\n\n")
        f.write(f"**Specification:** {spec_filename}\n\n")
        f.write(f"**Design Artifact:** {design_filename}\n\n")
        f.write("---\n\n")

        # Executive Summary
        f.write("## Executive Summary\n\n")
        f.write(f"This report presents the compliance analysis of **{design_filename}** ")
        f.write(f"against the requirements specified in **{spec_filename}**.\n\n")

        f.write("### Compliance Overview\n\n")
        f.write(f"| Metric | Count | Percentage |\n")
        f.write(f"|--------|-------|------------|\n")
        f.write(f"| Total Requirements | {total} | 100% |\n")
        f.write(f"| Compliant | {compliant} | {compliant/total*100:.1f}% |\n") if total > 0 else None
        f.write(f"| Partial | {partial} | {partial/total*100:.1f}% |\n") if total > 0 else None
        f.write(f"| Missing | {missing} | {missing/total*100:.1f}% |\n") if total > 0 else None
        f.write(f"| Unknown | {unknown} | {unknown/total*100:.1f}% |\n\n") if total > 0 else None

        f.write("### Risk Assessment\n\n")
        f.write(f"| Risk Level | Count |\n")
        f.write(f"|------------|-------|\n")
        f.write(f"| High | {high_risk} |\n")
        f.write(f"| Medium | {medium_risk} |\n")
        f.write(f"| Low | {low_risk} |\n\n")

        # Compliance rate indicator
        if compliance_rate >= 90:
            status = "GOOD"
            indicator = "+"
        elif compliance_rate >= 70:
            status = "NEEDS ATTENTION"
            indicator = "~"
        else:
            status = "CRITICAL"
            indicator = "!"

        f.write(f"**Overall Compliance Rate:** {compliance_rate:.1f}% [{status}]\n\n")
        f.write("---\n\n")

        # High Risk Items (Critical Findings)
        high_risk_results = [r for r in results if r.risk == RiskLevel.HIGH]
        if high_risk_results:
            f.write("## Critical Findings (High Risk)\n\n")
            f.write("The following requirements require immediate attention:\n\n")

            for result in high_risk_results:
                req = next((r for r in requirements if r.id == result.requirement_id), None)
                if req:
                    f.write(f"### {result.requirement_id}\n\n")
                    if req.clause:
                        f.write(f"**Clause:** {req.clause}\n\n")
                    f.write(f"**Requirement:** {req.description}\n\n")
                    f.write(f"**Verdict:** {result.verdict.value}\n\n")
                    f.write(f"**Risk Reason:** {result.risk_reason}\n\n")

                    if result.gaps:
                        f.write("**Gaps:**\n")
                        for gap in result.gaps:
                            f.write(f"- {gap.description}\n")
                            f.write(f"  - Suggested evidence: {gap.suggested_evidence}\n")
                            if gap.suggested_test:
                                f.write(f"  - Suggested test: {gap.suggested_test}\n")
                        f.write("\n")

            f.write("---\n\n")

        # Detailed Findings
        f.write("## Detailed Findings\n\n")

        # Group by verdict
        for verdict_type in [Verdict.MISSING, Verdict.PARTIAL, Verdict.UNKNOWN, Verdict.COMPLIANT]:
            verdict_results = [r for r in results if r.verdict == verdict_type]
            if not verdict_results:
                continue

            f.write(f"### {verdict_type.value} ({len(verdict_results)})\n\n")

            for result in verdict_results:
                req = next((r for r in requirements if r.id == result.requirement_id), None)
                if req:
                    f.write(f"#### {result.requirement_id}")
                    if req.clause:
                        f.write(f" (Clause {req.clause})")
                    f.write("\n\n")

                    f.write(f"**Type:** {req.type.value.upper()} | ")
                    f.write(f"**Priority:** {req.priority.value} | ")
                    f.write(f"**Risk:** {result.risk.value}\n\n")

                    f.write(f"> {req.description}\n\n")

                    f.write(f"**Rationale:** {result.rationale}\n\n")

                    if result.evidence:
                        f.write("**Evidence:**\n")
                        for ev in result.evidence[:3]:
                            f.write(f"- [{ev.source_location}]: \"{ev.quote[:150]}...\"\n")
                        f.write("\n")

                    if result.gaps:
                        f.write("**Gaps:**\n")
                        for gap in result.gaps:
                            f.write(f"- {gap.description}\n")
                        f.write("\n")

            f.write("\n")

        # Recommendations
        f.write("---\n\n")
        f.write("## Recommendations\n\n")

        if high_risk > 0:
            f.write(f"1. **Address {high_risk} high-risk items immediately** - ")
            f.write("These represent critical gaps in compliance.\n\n")

        if missing > 0:
            f.write(f"2. **Document or implement {missing} missing requirements** - ")
            f.write("No evidence was found for these specifications.\n\n")

        if partial > 0:
            f.write(f"3. **Complete documentation for {partial} partial items** - ")
            f.write("Additional evidence or implementation needed.\n\n")

        if unknown > 0:
            f.write(f"4. **Clarify {unknown} unknown status items** - ")
            f.write("Manual review recommended to determine compliance.\n\n")

        f.write("---\n\n")
        f.write("*Report generated by SpecGuard - Engineering Specification Compliance Checker*\n")

    logger.info(f"Compliance report generated: {output_path}")
    return output_path


def generate_requirements_json(
    requirements: list[Requirement],
    output_dir: Path,
    analysis_id: str
) -> Path:
    """Export requirements to JSON."""
    output_path = output_dir / f"requirements_{analysis_id}.json"

    data = [req.model_dump(mode="json") for req in requirements]

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)

    logger.info(f"Requirements JSON exported: {output_path}")
    return output_path


def generate_full_results_json(
    requirements: list[Requirement],
    results: list[ComplianceResult],
    spec_filename: str,
    design_filename: str,
    output_dir: Path,
    analysis_id: str
) -> Path:
    """Export full analysis results to JSON."""
    output_path = output_dir / f"full_results_{analysis_id}.json"

    # Calculate summary
    total = len(requirements)
    summary = AnalysisSummary(
        total_requirements=total,
        compliant_count=sum(1 for r in results if r.verdict == Verdict.COMPLIANT),
        partial_count=sum(1 for r in results if r.verdict == Verdict.PARTIAL),
        missing_count=sum(1 for r in results if r.verdict == Verdict.MISSING),
        unknown_count=sum(1 for r in results if r.verdict == Verdict.UNKNOWN),
        high_risk_count=sum(1 for r in results if r.risk == RiskLevel.HIGH),
        medium_risk_count=sum(1 for r in results if r.risk == RiskLevel.MEDIUM),
        low_risk_count=sum(1 for r in results if r.risk == RiskLevel.LOW),
    )

    data = {
        "analysis_id": analysis_id,
        "spec_filename": spec_filename,
        "design_filename": design_filename,
        "generated_at": datetime.utcnow().isoformat(),
        "summary": summary.model_dump(mode="json"),
        "requirements": [req.model_dump(mode="json") for req in requirements],
        "results": [res.model_dump(mode="json") for res in results]
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)

    logger.info(f"Full results JSON exported: {output_path}")
    return output_path


def generate_xlsx_matrix(
    requirements: list[Requirement],
    results: list[ComplianceResult],
    output_dir: Path,
    analysis_id: str
) -> Optional[Path]:
    """
    Generate traceability matrix in XLSX format (optional).

    Requires openpyxl to be installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.warning("openpyxl not installed, skipping XLSX generation")
        return None

    result_lookup = {r.requirement_id: r for r in results}
    output_path = output_dir / f"traceability_matrix_{analysis_id}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Traceability Matrix"

    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    # Verdict colors
    verdict_colors = {
        "Compliant": "C6EFCE",
        "Partial": "FFEB9C",
        "Missing": "FFC7CE",
        "Unknown": "D9D9D9",
    }

    # Headers
    headers = [
        "Requirement ID", "Clause", "Requirement", "Type", "Priority",
        "Verdict", "Evidence", "Gaps", "Risk", "Risk Reason"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row, req in enumerate(requirements, 2):
        result = result_lookup.get(req.id)

        ws.cell(row=row, column=1, value=req.id)
        ws.cell(row=row, column=2, value=req.clause or "")
        ws.cell(row=row, column=3, value=req.description[:200])
        ws.cell(row=row, column=4, value=req.type.value)
        ws.cell(row=row, column=5, value=req.priority.value)

        if result:
            verdict_cell = ws.cell(row=row, column=6, value=result.verdict.value)
            color = verdict_colors.get(result.verdict.value, "FFFFFF")
            verdict_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            evidence_refs = "; ".join(str(ev.source_location) for ev in result.evidence[:3])
            ws.cell(row=row, column=7, value=evidence_refs or "No evidence")

            gaps = "; ".join(gap.description for gap in result.gaps)
            ws.cell(row=row, column=8, value=gaps)

            ws.cell(row=row, column=9, value=result.risk.value)
            ws.cell(row=row, column=10, value=result.risk_reason)
        else:
            ws.cell(row=row, column=6, value="Not evaluated")

    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 40
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 40

    wb.save(output_path)
    logger.info(f"XLSX traceability matrix generated: {output_path}")
    return output_path
